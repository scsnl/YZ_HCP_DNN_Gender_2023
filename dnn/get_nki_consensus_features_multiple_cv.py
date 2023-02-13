import math
import json
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from nilearn import plotting,image,surface,datasets
from utilityFunctions import *
import collections

bn_atlas_file = 'PROJECT_DIR/scripts/features/bnatlas_tree.json'

parent_substitutions={
    'Amyg, Amygdala': 'Middle Temporal Gyrus',
    'CG, Cingulate Gyrus': 'Cingulate Gyrus',
    'Cun, Cuneus': 'PCC, Precuneus',
    'Frontal Lobe': 'Prefrontal Cortex',
    'FuG, Fusiform Gyrus': 'Inferior Temporal Gyrus',
    'Hipp, Hippocampus' : 'Middle Temporal Gyrus',
    'IFG, Inferior Frontal Gyrus': 'Prefrontal Cortex',
    'INS, Insular Gyrus' : 'Prefrontal Cortex',
    'IPL, Inferior Parietal Lobule': 'Inferior Parietal Lobe',
    'ITG, Inferior Temporal Gyrus': 'Inferior Temporal Gyrus',
    'Insular Lobe':'Prefrontal Cortex',
    'Limbic Lobe' : 'Middle Temporal Lobe',
    'MFG, Middle Frontal Gyrus' : 'Prefrontal Cortex',
    'MTG, Middle Temporal Gyrus' : 'Middle Temporal Gyrus',
    'OcG, Occipital Gyrus' : 'Occipital Gyrus',
    'Occipital Lobe':'Occipital Lobe',
    'OrG, Orbital Gyrus' : 'Prefrontal Cortex',
    'PCL,Paracentral Lobule' : 'Paracentral Lobule',
    'Parietal Lobe' : 'Parietal Lobe',
    'Pcun, Precuneus' : 'PCC, Precuneus',
    'PhG, Parahippocampal Gyrus' : 'Middle Temporal Gyrus',
    'PoG, Postcentral Gyrus' : 'Postcentral Gyrus',
    'PrG, Precentral Gyrus' : 'Precentral Gyrus',
    'Psts, Posterior Superior Temporal Sulcus' : 'Superior Temporal Gyrus',
    'SFG, Superior Frontal Gyrus' : 'Prefrontal Cortex',
    'SPL, Superior Parietal Lobule' : 'Superior Parietal Lobule',
    'STG, Superior Temporal Gyrus' : 'Superior Temporal Gyrus',
    'Str, Striatum' : 'Striatum',
    'Subcortical Nuclei' : 'Subcortical Nuclei',
    'Temporal Lobe' : 'Temporal Lobe',
    'Tha, Thalamus' : 'Thalamus',
              }

def write_feature_consensus_excel_file(occurrences,bn_atlas_file,excel_file,group,percentile):

    with open(bn_atlas_file) as f:
        bn_atlas=json.load(f)

    if not os.path.exists(excel_file):
        from openpyxl import Workbook
        wb = Workbook()
        ws1= wb.active
        ws1.title="%s_%02d"%(group,percentile)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws1 = wb.create_sheet(title="%s_%02d"%(group,percentile))

    ws1.append(["Region ID","Gyrus","Description","Region Alias","(ID) Region Label","Count"])

    feature_data=[]
    for feature in occurrences:
        # print(feature, type(feature), occurrences[feature], type(occurrences[feature]))
        featureID = str(feature + 1) # ROI/feature ID
        for idx,region in enumerate(bn_atlas):
            if region['id'] == featureID:
                # print("feature: {}, frequency: {}\n".format(featureID, occurrences[feature]))
                feature_data.append([featureID, parent_substitutions[region['parent']],region['text'],region['data']['alias'],
                           "(%s), %s"%(featureID, region['text']),"%01d"%(occurrences[feature])])

    for feature in feature_data:
        ws1.append(feature)

    tab = Table(displayName="group%s_percentile%02d"%(group,percentile), ref="A1:F%d"%(len(occurrences)+1))
    style = TableStyleInfo(name="TableStyleLight15", showFirstColumn=False,
                       showLastColumn=True, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws1.add_table(tab)

    table_cells = ws1['A1':'F%d'%(len(occurrences)+1)]

    cols=['A','B','C','D','E','F']

    for idx,col in enumerate(ws1.columns):
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws1.column_dimensions[cols[idx]].width = math.ceil(adjusted_width)

    for row in table_cells:
        for cell in row:
            cell.alignment = Alignment(horizontal="center",vertical='center',wrapText=True)
    wb.save(filename = excel_file)

    return feature_data

def save_feature_consensus_nifti(occurrences, output_nii_path, group, site, percentile):
    bn_nifti = 'PROJECT_DIR/scripts/features/BN_Atlas_246_2mm.nii'

    atlas_volume = image.load_img(bn_nifti)
    roi_nifti = image.math_img('img-img', img=atlas_volume)
    img_data = atlas_volume.get_data()

    for feature in occurrences:
        # print(feature, type(feature), occurrences[feature], type(occurrences[feature]))
        roi_idx = np.where(img_data == feature + 1, (occurrences[feature]*1.0), 0)
        roi_img = image.new_img_like(roi_nifti, roi_idx)
        roi_nifti = image.math_img('img1+img2', img1=roi_nifti, img2=roi_img)

    output_nii_file = os.path.join(output_nii_path,
                               'bn_features_consensus_group_%s_site_%s_percentile_%02d.nii.gz' % (group, site, percentile))
    print(output_nii_file)
    roi_nifti.to_filename(output_nii_file)


if __name__ == '__main__':

    # which sessions models were trained on
    model_sessions_list = ['LR_S1', 'RL_S1']
    # which normz non-windowed entire dataset were used for generating feature attribution
    test_sessions_list = ['645', '645']
    test_sessions_alias_list = ['nki_645', 'nki_645']
    num_5folds = 100

    output_path = 'PROJECT_DIR/results/restfmri/dnn/feature_excel_files/'
    if not os.path.exists(output_path):
        os.makedirs(output_path)

    output_nii_path = 'PROJECT_DIR/results/restfmri/dnn/niis/consensus_multiple_cv/'
    if not os.path.exists(output_nii_path):
        os.makedirs(output_nii_path)

    for ii in range(2):
        model_session = model_sessions_list[ii]
        test_session = test_sessions_list[ii]
        test_session_alias = test_sessions_alias_list[ii]

        output_file = output_path + 'hcp_model_' + model_session + '_test_' + test_session_alias + '_consensus_multiple_cv.xlsx'
        print("output file: {}".format(output_file))

        groups = {'male': 0, 'female': 1}

        for percentile in [80, 85, 90, 95]: # for each percentile
            for group in groups.keys(): # for each gender
                # get top features of each fold and count
                # the number of appearance of each feature across folds
                # and write features to excel
                feature_data = []
                data_path = 'PROJECT_DIR/results/restfmri/dnn/attributions/multiple_cv/'
                site = 'hcp_model_' + model_session + '_test_nki_' + test_session

                for jj in range(num_5folds):
                    print("the {} 5folds".format(jj))

                    for m in range(5):
                        data_file = data_path + 'ff' + str(jj) + '_hcp_model_' + model_session + '_index_' + str(m) + '_test_nki_' + test_session + '.npz'
                        print(jj, m, group, percentile)
                        feature_idcs, features = determine_features(data_file, groups[group], percentile)
                        feature_data.append(feature_idcs)

                print(len(feature_data))
                # get all top ROI/feature IDs from all folds and flatten all lists into one list
                feature_data_flatten = [ft for sublist in feature_data for ft in sublist]
                print(len(feature_data_flatten))
                # count the occurrence of each top feature
                occurrences = collections.Counter(feature_data_flatten)
                print(occurrences)
                # occurerences is a dictionary with keys = ROI/feature IDs, and values = #occurrence
                # write results to excel
                write_feature_consensus_excel_file(occurrences,bn_atlas_file,output_file,group,percentile)
                # write results to nii for plot
                save_feature_consensus_nifti(occurrences, output_nii_path, group, site, percentile)