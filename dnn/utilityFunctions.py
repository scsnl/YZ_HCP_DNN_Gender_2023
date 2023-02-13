import numpy as np
import os
from captum.attr import IntegratedGradients
from captum.attr import DeepLift
# from openpyxl.worksheet.table import Table, TableStyleInfo
# from openpyxl.styles import Alignment
from nilearn import plotting,image,surface,datasets
from scipy.spatial.distance import squareform, pdist

def reshapeData(data):
    no_subjs, no_ts, no_channels = data.shape
    # Reshape data to no_subjs, no_channels, no_ts
    data_reshape = np.empty((no_subjs, no_channels, no_ts))
    for subj in np.arange(no_subjs):
        x_subj = data[subj, :, :]
        x_subj = np.transpose(x_subj)
        data_reshape[subj, :, :] = x_subj
    return data_reshape

def prepare_data_sliding_window(data, labels,window_size, step):
    ''' Function to create windowed data'''
    Nsubjs,N,Nchannels = data.shape
    width = np.int(np.floor(window_size / 2.0))
    labels_window = []
    window_data_list=[]
    for subj in np.arange(Nsubjs):
        #print("subject = ",subj)
        for k in range(width, N - width - 1, step):
            x = data[subj,k - width: k + width,:]
            x = np.expand_dims(x,axis=0)
            window_data_list.append(x)
            labels_window.append(labels[subj])
    window_data = np.vstack(window_data_list)
    return (window_data,labels_window)


def write_excel_file(accuracy, precision, recall, f1, excel_file, model_ss, test_ss):
    if not os.path.exists(excel_file): # excel file name
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "trained_%s_tested_%s" % (model_ss, test_ss) # excel sheet name
    else:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws1 = wb.create_sheet(title="trained_%s_tested_%s" % (model_ss, test_ss))

    ws1.append(["Session", "Fold Number", "Accuracy", "Precision", "Recall", "F1-score"])

    for idx in range(len(accuracy)):
        if idx == 0:
            ws1.append([test_ss, "%01d" % (idx+1), "%02.02f" % accuracy[idx],
                       "%02.02f" % (precision[idx]*100), "%02.02f" % (recall[idx]*100),
                        "%02.02f" % (f1[idx]*100)])
        else:
            ws1.append(["", "%01d" % (idx + 1), "%.02f" % accuracy[idx],
                       "%.02f" % (precision[idx]*100), "%.02f" % (recall[idx]*100),
                        "%.02f" % (f1[idx]*100)])

    ws1.append(["", "Avg (Std)",
               "%.02f (%.02f)" % (np.mean(accuracy), np.std(accuracy)),
               "%.02f (%.02f)" % (np.mean(precision)*100, np.std(precision)*100),
               "%.02f (%.02f)" % (np.mean(recall)*100, np.std(recall)*100),
               "%.02f (%.02f)" % (np.mean(f1)*100, np.std(f1)*100)])

    wb.save(filename=excel_file)


# target should be 0 for male and 1 for female as we encoded male as 0 and female as 1
def getInputAttributions(model, input_tensor,target):
    ig = IntegratedGradients(model)
    input_tensor.requires_grad_()
    attr, delta = ig.attribute(input_tensor, target=target, return_convergence_delta=True)
    attr = attr.cpu().detach().numpy()
   # attr = attr.detach().numpy()
    return attr

# target should be 0 for male and 1 for female as we encoded male as 0 and female as 1
def getInputAttributions_DeepLift(model, input_tensor,target):
    ig = DeepLift(model,multiply_by_inputs=True)
    input_tensor.requires_grad_()
    attr, delta = ig.attribute(input_tensor, target=target, return_convergence_delta=True)
    attr = attr.cpu().detach().numpy()
   # attr = attr.detach().numpy()
    return attr

def determine_features(data_file,group_label,percentile):

    data = np.load(data_file)
    # for k in data.files:
    #     print(k)
    # print(data['features'].shape) # num_sub x num_roi x num_time_point
    # get data for subjects within a specific group
    group_features = data['features'][np.where(data['labels'] == group_label)]
    medians = np.median(group_features, axis=2) # get medians across time points
    mean_across_subj = np.mean(np.abs(medians), axis=0) # average abs medians across subjects
    # the remaining dimension is num_roi (246)
    # print("mean_across_subj shape is {}".format(mean_across_subj.shape))
    percentiles = np.where(np.abs(mean_across_subj) >= np.percentile(np.abs(mean_across_subj),percentile))
    features_idcs = percentiles[0] # includes all indices (rois) at which position the values are above the cutoff
    features = mean_across_subj # feature scores (averaged across subjects)

    # features_idcs element + 1 = feature ID
    # features = feature attribution weights
    return features_idcs,features


def save_nifti(features_idcs, features, output_dir, group, site, percentile):
    bn_nifti = 'PROJECT_DIR/scripts/features/BN_Atlas_246_2mm.nii'

    atlas_volume = image.load_img(bn_nifti)
    roi_nifti = image.math_img('img-img', img=atlas_volume)
    img_data = atlas_volume.get_data()

    for idx in features_idcs:
        roi_idx = np.where(img_data == idx + 1, features[idx], 0)
        roi_img = image.new_img_like(roi_nifti, roi_idx)
        roi_nifti = image.math_img('img1+img2', img1=roi_nifti, img2=roi_img)

    output_file = os.path.join(output_dir,
                               'bn_features_group_%s_site_%s_percentile_%02d.nii.gz' % (group, site, percentile))
    print(output_file)
    roi_nifti.to_filename(output_file)

def save_indiv_nifti(subjid, features_idcs, features, output_dir, group, site, percentile):
    bn_nifti = 'PROJECT_DIR/scripts/features/BN_Atlas_246_2mm.nii'

    atlas_volume = image.load_img(bn_nifti)
    roi_nifti = image.math_img('img-img', img=atlas_volume)
    img_data = atlas_volume.get_data()

    for idx in features_idcs:
        roi_idx = np.where(img_data == idx + 1, features[idx], 0)
        roi_img = image.new_img_like(roi_nifti, roi_idx)
        roi_nifti = image.math_img('img1+img2', img1=roi_nifti, img2=roi_img)

    output_file = os.path.join(output_dir,
                               'bn_features_%s_%s_%s_percentile_%02d.nii.gz' % (group, subjid, site, percentile))
    print(output_file)
    roi_nifti.to_filename(output_file)


def distcorr2(X, Y):
    # X = np.atleast_1d(X)
    # Y = np.atleast_1d(Y)
    # if np.prod(X.shape) == len(X):
    #     X = X[:, None]
    # if np.prod(Y.shape) == len(Y):
    #     Y = Y[:, None]
    # X = np.atleast_2d(X)
    # Y = np.atleast_2d(Y)
    n = X.shape[0]
    # if Y.shape[0] != X.shape[0]:
    #     raise ValueError('Number of samples must match')
    a = squareform(pdist(X))
    b = squareform(pdist(Y))
    A = a - a.mean(axis=0)[None, :] - a.mean(axis=1)[:, None] + a.mean()
    B = b - b.mean(axis=0)[None, :] - b.mean(axis=1)[:, None] + b.mean()
    dcov2_xy = (A * B).sum() / float(n * n)
    dcov2_xx = (A * A).sum() / float(n * n)
    dcov2_yy = (B * B).sum() / float(n * n)
    dcor = np.sqrt(dcov2_xy) / np.sqrt(np.sqrt(dcov2_xx) * np.sqrt(dcov2_yy))
    return dcor*dcor
